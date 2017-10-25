from openpyxl import load_workbook
from twilio.rest import TwilioRestClient
import datetime, schedule, time

today = False

client = TwilioRestClient("AC95f0cc1ed287910527349014305fe1f4", "b127c2204e5a2a8948b21585e59a4428")

# Period timings
times={'P1':'07:55','P2':'08:35','P3':'09:15','P4':'09:55','P5':'10:45','P6':'12:45','P7':'12:10','P8':'01:10','P9':'01:50'}

# open Excel file with Teacher phone numbers
n = {}
nums = load_workbook(filename = 'Numbers.xlsx')
sheet_ranges = nums['Sheet1']
names = sheet_ranges['A']
numbers = sheet_ranges['B']
for i in range(len(names)):
	n[str(names[i].value)]="+"+str(numbers[i].value)

# open Excel file with substitution info
SubLogs = load_workbook(filename = 'Log1.xlsx')
sheet_ranges = SubLogs['Sheet1']
periods = sheet_ranges['D']
teachers = sheet_ranges['F']
classroom = sheet_ranges['E']
dat = sheet_ranges['A']
def check():

	global today

	# Current time
	t = str(datetime.datetime.now().time())
	
	# Today's date
	d = str(datetime.datetime.now())[:10]
	d = datetime.datetime.strptime(d,'%Y-%m-%d').strftime('%Y-%d-%m')

	substitutes = []
	for i in range(len(periods)):
		if str(dat[i].value)[:10] == d:
			today = True
		elif str(dat[i].value)[:10] != None:
			today = False
		if periods[i].value != None:
			if periods[i].value in times:
				if t[:5] == times[periods[i].value] and today:
					substitutes.append([str(teachers[i].value),str(classroom[i].value)])
	for i in substitutes:
		if i[0] in n:
			print "Sending message"
			message="Reminder: You have substitution in class "+i[1]
			client.messages.create(to=n[i[0]], from_="+12013895808", body=message)
check()
schedule.every(1).minutes.do(check)

while True:
    schedule.run_pending()
    time.sleep(1)